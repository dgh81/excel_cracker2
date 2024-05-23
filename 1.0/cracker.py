from openpyxl import load_workbook

def crack(dir_file_path):
    try:
        if dir_file_path.endswith("xlsm"):
            wb = load_workbook(filename=dir_file_path, read_only=False, keep_vba=True)
        else:
            wb = load_workbook(filename=dir_file_path, read_only=False)

        for ws in wb.worksheets:
            ws.protection.sheet = False

        # print(wb.security.lockStructure)
        try:
            if wb.security.lockStructure == True:
                wb.security.lockStructure = False
        except Exception as e:
            print("Exception: ", e)

        wb.save(dir_file_path)
        wb.close()
        
    except Exception as e:
        print("Exception: ",e)
        return False
    return True

